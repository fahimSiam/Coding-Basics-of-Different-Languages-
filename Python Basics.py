-*- coding: utf-8 -*-
"""
Created on Mon Apr 20 17:36:48 2020

@author: Asus
"""

price = 10
price = 20
print(price)
print(price*10)

rating = 4.9

name = 'mosh'
is_published = True
name= 'paul'
age= 23
is_new= True

name= input('what is your name?')
print('hi '+ name)
name=input('what is your name?')
color= input('what is your favourite color?')
print(name+' likes '+color)

birth_year=input('give your birth year: ')
print(type(birth_year))
age= 2020- int(birth_year)
print(type(age))
print(age)

weight_pound=input( 'your weight in pounds? Ans:' )
weight_kg=( int(weight_pound)/2)-(int(weight_pound)/20)
print('your weight in kg is: '+str(weight_kg))

print('''
      so let's see,
      what happens
      when we put triple qoutes
''')

line='sami ekta bolod'
print(line[0:4])

line='sami ekta bolod'
print(line[0:])

line='sami ekta bolod'
print(line[-6:])

line='sami ekta bolod'
another= line[0:4]
print(another[:])

fisrt= 'fahim'
last= 'siam'
msg= f'{fisrt} [{last}] is a coder'
print(msg)

course= 'Python for Beginners'
print(len(course)) #calculate the number of characters
course.upper() #method bcoz only applied to strings and print function.
print (course.upper())
print(course.lower())
print(course.find('y')) #index
print(course.replace('Beginners','Absoloute Beginners'))
print('Python' in course) #boolian in-operator
#arithmetics
print(10/3)#the full value
print(10//3)#only the int
print(10%3)#remainder
print(10**3)#power

x= 10
x = x+3
x+=3
print(x)

x= -2.9
print(round(x))
print(abs(round(x)))

import math
print(math.ceil(2.8))
print(math.floor(2.8))
#seach python 3 math modules
#IF ELIF ELSE STATEMENTS
is_hot = False
is_cold = False
if is_hot:
    print("it's a hot day")
    print("drink plenty of water")
elif is_cold:
    print("it's a cold day")
    print("wear warm clothes")
else: #shift+tab
    print("it's a beautiful day")

print ("enjoy your day")
#shift+tab
     
price = 1000000
good_credit = True
if good_credit:
    price= price* 0.1
else:
    price= price*20
print(f"down payment: {price} dollars")

#double conditions
has_high_income = True
has_good_income = True
if has_high_income and has_good_income:
    print('eligable for loan')


has_high_income = True
has_good_income = True
if has_high_income and not has_good_income:
    print('eligable for loan')
else:
    print("not eligable")

temp=30
if temp != 30:
    print("it's a hot day")
else: 
    print("it's not a hot day")

# NAME SIZE CHECK
name=(input("What's your name?"))
if 3 >= (len(name)):
    print('name must be atleast 3 characters')
elif 50 <= (len(name)):
    print('name can be maximum of 50 characters')
else: print('name looks good!')

weight= (input('what is your weight?'))
metric= (input('(L)bs or (k)bs?'))
if metric.lower() == 'l':
    weight = int(weight) *0.45
    print(f"you are {weight} kgs")
elif metric.lower()=='k':
    weight= int(weight)/0.45
    print(f"your weight is pound is {weight}")

#while loop
i = 1
while i<=5:
    print('*'*i)
    i+=1
print('ends')

#guessing game
guess= 3
attempt = -1
count=0
while count<3:
    attempt= int(input('what is your guess?'))
    count +=1
    if guess == attempt:
        print('your guess is correct')
        break
if attempt != guess:
    print('you have no attempts left')

#car game
count =0
boolean = False
while count <50:
    command= input('what do u want your car to do?')
    count +=1
    if command.lower() == 'help':
        print(''' 
type "start" to start your car
type "stop" to stop your car
type "quit" to exit game''')
    elif command.lower() == 'start' and not boolean:
        boolean = True
        print('your car has started')
    elif command.lower() == 'start' and boolean:
        print('car has already started')
    elif command.lower() == 'stop' and boolean :
        boolean = False
        print('your car has stopped')
    elif command.lower() == 'stop' and not boolean:
        print('car is already stopped')
    elif command.lower() == 'quit':
        print('nice timepass korlen bhai')
        break
    else: 'esob ki bolen bhai?'
print('onek khelsen, dhonnobad')
####### if not boolean: (not operator o use kora jae uporer tae)

# FOR LOOP

for item in 'python':
    print (item)

for item in ['siam','faiza','moo','kibria','muizz','galib','trina']:
    print(item)

for item in range(10):
    print(item)

for item in range(5,20,2):
    print(item)
    
total=0
for prices in [10,20,30]:
    total+=prices
print(total)

for x in range(4):
    for y in range(4):
        print(f'({x},{y})')
        
num=[5,2,5,2,2]
for times in num:
    output=''
    for count in range(times):
        output+='x'
    print(output)
    
name = ['siam','faiza','moo','kibria','muizz','galib','trina']
print(name[0])



#check which is the largest
numbers=[3,6,4,7,9]
largest = numbers[0]
for number in numbers:
    if number >largest:
        largest= number
print(largest)

#2d array
matrix =[
        [1,2,3],
        [4,5,6],
        [7,8,9]
        ]
print(matrix[0][0])
for row in matrix:
    for item in row:
        print(item)

numbers=[5,12,45,34,2,1,5]
numbers.append(20)#add at the end
numbers.insert(0,10)# add in specific
numbers.remove(45)
print(numbers)
numbers.pop()#delete last one
print(numbers.index(5))
print(50 in numbers)
print(numbers)
print(numbers.count(5))
numbers.sort()#sorting
numbers.reverse()#decending
numbers2=numbers.copy()
numbers.append(69)
print(numbers)
print(numbers2)
numbers.clear()
print(numbers)

#remove dublicate
numbers=[1,2,3,4,5,6,3,4,5,5,6,21]
numbers.sort()
for check in numbers:
    if numbers[check]==numbers[check-1]:
        numbers.remove(check)
print(numbers)

numbers=[1,2,3,4,5,6,3,4,5,5,6,21]
uniques=[]
for number in numbers:
    if number not in uniques:
        uniques.append(number)
print(uniques)

#topples and UNPACKING
numbers=(1,2,3,4,5)
print(numbers[2])

x,y,z,a,b= numbers
print(x)

 #dictionary
 customer={
         "name": "fahim siam",
         "age": 30,
         "is_varified":True
}
 customer["name"]= "jack sparrow"
 customer["id"]= "20141040"
 print(customer["id"])
print(customer.get("birthday", "jan 19 1997"))

#phone number convert
phone = input("phone :")
digits_mapping ={
        '1': 'One',
        '2': 'two',
        '3': 'three',
        '4': 'four',
        '5': 'five',
        '6': 'six',
        '7':'seven',
        '8':'eight',
        '9':'nine',
        '0':'zero'
 }
output=" "
for ch in phone:
   output+=digits_mapping.get(ch,'!')+' '
print(output)

#make emojis
message =input(">")
words=message.split(' ') 
print(words)
emojis={
        ":)": "😄",
        ":(" : "😥"
        }
output= ""
for word in words:
    output+= emojis.get(word, word) + " "
print(output)

#function
def greet_user(first_name, last_name):#paramiter
    print(f'hi there {first_name} {last_name}')
    print ('welcome aboard')
print('start')
greet_user('John',last_name='smith')
greet_user(last_name='fahim' ,first_name='siam')#argument
print('end')

def square(number):
    return number*number
result=square(3)
print(result)
print(square(4))

#function emoji
def emoji(message):
    words=message.split(' ')
    emojis={
            ":)":"😄",
            ":(":"😥",
            "<3":"❣"
            
            }
    output=""
    for word in words:
        output+=emojis.get(word , word) + " "
    return output
message=input('what is the sentence: ')
print(emoji(message))

#error handle
try:
    age=int(input('age: '))
    income= 20000
    risk=income/age
    print(age)
except ValueError:
    print('Invalid value')    
except ZeroDivisionError:
    print('zero disos ken')

#classes
class Point:
    def __init__(self,x,y):
        self.x = x
        self.y = y
        
    def move(self):
        print=("move")
    def draw(self):
        print('draw')
point1=Point(10,20)
#point1.x=10
#point1.y=20
print(point1.x)
point1.draw()

class Person:
    def __init__(self, name):
        self.name= name
    def talk(self):
        print(f'hi, i am {self.name}')
        
        
john= Person('john smith')
print(john.name)
john.talk()
bob=Person('bob smith')
bob.talk()    

class Employee:
    def __init__(self, first,last,pay):
        self.first= first
        self.last=last
        self.pay=pay
        self.email=first+ '.'+last+'@company.com'
    def fullname(self):
        return '{} {}' .format(self.first,self.last)
    def apply_raise(self):
        self.pay= int(self.pay *1.04)
    @classmethod
    def set_raise_amt(cls,amount):
        cls.raise_amt=amount
emp_1= Employee('Coray','Schafer', 50000)
emp_2= Employee('Test','User',60000)
print(emp_2.fullname())
Employee.set_raise_amt(1.05)
print(Employee.raise_amt)

#inheritance
class Mammal:
    def walk(self):
        print('walk')
     
        
class Dog(Mammal):
    def bark(self):
        print('bark')


class Cat(Mammal):
    def be_annoying(self):
        print('annoying')

dog1=Dog()
dog1.walk()
cat1=Cat()
cat1.be_annoying()
    
#modules
import converters
print(converters.kg_to_lbs(70))

import ecommerce.shipping
ecommerce.shipping.calc_shipping()

import random
for i in range(3):
    print(random.random())
    print(random.randint(10,20))

import random
line=' '
for hates in range(10):
    members=['faiza','siam','moo','muizz','kibria','galib','trina']
    luiccha=random.choice(members)
    lucchi=random.choice(members)
    line+=f"{luiccha} hates {lucchi}   "
print(line)

#random dice
import random
class  Dice:
    def roll(self, one, two):
        print(one)
        print(two)
        #numbers=(1,2,3,4,5,6)
        num1=random.randint(1,6)
        num2=random.randint(1,6)
        return num1,num2  #topple
dice=Dice()
print(dice.roll(1,2))

from pathlib import Path
#absolute path(pura directory) c:\progra
#relative path(ekta theke arektar)
path=Path("ecommerce")
print(path.exists())
path=Path("message")
#print(path.mkdir())#make directory
print(path.rmdir())#remove directory


from pathlib import Path
path=Path()
for file in path.glob('*.py'):
    print(file)
#pipy.org



import openpyxl as xl
wb=xl.load_workbook('transactions.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
cell=sheet.cell(1,1)
print(cell.value)
print(sheet.max_row)#num of rows

for row in range(2, sheet.max_row+1)
    print(row)
    cell= sheet.cell(row,3)
    print(cell.value)


import openpyxl as xl
from openpyxl.chart import BarChart,Reference#for the bar (chart module theke)
def process_workbook(filemane):
    wb=xl.load_workbook(filename)
    Sheet=wb['Sheet1']#access sheet
    cell = Sheet['a1']#access cell
    cell=Sheet.cell(1,1)#same
    print(cell.value)
    print(Sheet.max_row)#obj sheet er num of rows
    #corrected_price= Sheet.cell(1,1)
    #corrected_price.value=corrected_price
    for row in range(2, Sheet.max_row+1):# range doesn't include the last num and 1st ta row name
        cell= Sheet.cell(row,3)
        print(cell.value)
        corrected_price=cell.value*0.9
        corrected_price_cell=Sheet.cell(row,4)
        corrected_price_cell.value=corrected_price
    values=Reference(Sheet,
                     min_row=2,
                     max_row=Sheet.max_row,
                     min_col=4,
                     max_col=4)
    chart =BarChart()
    chart.add_data(values)
    Sheet.add_chart(chart, 'e2')#e2 is which cell u wanna put the chart
    wb.save(filename)


#JUPYTER NOTEBOOK
#KAGGLE.COM FOR SAMPLE csv file
#import pandas as pd
df=pd.read_csv('vgsales.csv')
df.shape
df.describe()
df.values
# shows the attributes...count is how many data it has....mean=average...
# double D to delete in command mode
# B for new cell A for above
#ctrl+/ for comment

X=op.drop(columns=['rating'])
X

Y=op['rating']
Y


import pandas as pd
from sklearn.tree import DecisionTreeClassifier
op=pd.read_csv('anime.csv')
X=op.drop(columns=['members'])
y=op['members']

model = DecisionTreeClassifier()
model.fit(X, y)
predictions = model.predict([696969, 'two piece' , 'hentai','OVA',69])
predictions

import pandas as pd
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score

op=pd.read_csv('anime.csv')
X=op.drop(columns=['members'])
y=op['members']
X_train,X_test, y_train, y_test = train_test_split(X,y, test_size=0.2)

model = DecisionTreeClassifier()
model.fit(X_train, y_train)
predictions = model.predict(X_test)

accuracy_score
# predictions



import pandas as pd
from sklearn.tree import DecisionTreeClasssifier
form sklearn import tree

op=pd.read_csv('anime.csv')
X=op.drop(columns=['members','name','genre','type'])
y=op['members']

model= DecisionTreeClasssifier()
model.fit(X,y)

tree.export_graphviz(model, out_file='anime-predictor.dot',
                    feature_name=['anime_id','name','genre','type','episodes','members'],
                    class_names=sorted(y.unique()),
                     label='all',
                     rounded=True
                     filled=True
                    )























































